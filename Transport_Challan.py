from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import json, os

COUNTER_FILE = "file_counter.json"

# Initialize the counter file if it doesn't exist
def initialize_counter():
    if not os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, "w") as file:
            json.dump({"counter": 1}, file)

# Get the current counter value
def get_next_counter():
    with open(COUNTER_FILE, "r") as file:
        data = json.load(file)
        counter = data["counter"]
        data["counter"] += 1

    # Update the counter file
    with open(COUNTER_FILE, "w") as file:
        json.dump(data, file)

    return counter


# File to store company data
COMPANY_DATA_FILE = r"Data\company_data.json"
TRANSPORT_DATA_FILE = r"Data\transport_data.json"

# Load existing company data from JSON file
def load_data_file(file_path):
    if not os.path.exists(file_path):
        with open(file_path, 'w') as file:
            json.dump({}, file)  # Create a file with an empty JSON object

    try:
        with open(file_path, 'r') as file:
            if os.stat(file_path).st_size == 0:  # Check if the file is empty
                with open(file_path, 'w') as file:
                    json.dump({}, file)  # Write an empty JSON object if the file is empty
            file.seek(0)  # Move the file pointer to the beginning of the file
            return json.load(file)  # Load the JSON content
    except FileNotFoundError:
        return {}  # Return an empty dictionary if the file is not found
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in file {file_path}. Returning empty data.")
        return {}  # Return an empty dictionary if the JSON is invalid

# Save company data to JSON file
def save_data_file(file_path, data):
    """Save data to a JSON file."""
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

# Apply border to all used cells
def apply_borders(ws):
    """Apply thin borders to all cells in the specified range."""
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=1, max_row=32, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border


# Create the transport challan format
def format_ws(ws):
    bold_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # increase the height all cell
    for i in range(1, 33):
        ws.row_dimensions[i].height = 20

    # setting the size of cell f
    ws.column_dimensions['F'].width = 4

    # Transport Challan Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "TRANSPORt CHALLAN"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Supplier Name
    ws.merge_cells('A2:J2')
    ws['A2'] = "Supplier Details"
    ws['A2'].font = Font(bold=True, size=18)
    ws['A2'].alignment = center_align

    # Supplier Details
    ws.merge_cells('A3:J3')
    ws['A3'] = "Supplier Address(1)"
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align

    ws.merge_cells('A4:J4')
    ws['A4'] = "Supplier Address(2)"
    ws['A4'].font = bold_font
    ws['A4'].alignment = center_align

    ws.merge_cells('A5:J5')
    ws['A5'] = "GST No."
    ws['A5'].font = bold_font
    ws['A5'].alignment = center_align
    
    # Contact No.
    ws.merge_cells('A6:G6')
    ws.merge_cells('H6:J6')
    ws['H6'] = "Contact No."
    ws['H6'].font = bold_font
    ws['H6'].alignment = right_align

    # Transport 
    ws.merge_cells('A7:C7')
    ws['A7'] = "Transport:"
    ws['A7'].font = bold_font
    ws['A7'].alignment = center_align
    ws.merge_cells('D7:J7')
    ws['D7'] = "Transport Name"
    ws['D7'].font = bold_font
    ws['D7'].alignment = left_align

    ws.merge_cells('A8:J8')

    # Customer Detales
    ws.merge_cells('A9:D9')
    ws['A9'] = "Customer Name "
    ws['A9'].font = bold_font
    ws['A9'].alignment = center_align
    ws.merge_cells('A10:D10')
    ws['A10'] = "Customer Address (Contact No.)"
    ws['A10'].font = bold_font
    ws['A10'].alignment = center_align
    ws.merge_cells('A11:D11')
    ws['A11'] = "GST No."
    ws['A11'].font = bold_font
    ws['A11'].alignment = center_align

    ws.merge_cells('E9:F11')

    # Challan Number
    ws.merge_cells('G9:H9')
    ws['G9'] = "ch no."
    ws['G9'].font = bold_font
    ws['G9'].alignment = center_align
    ws.merge_cells("I9:J9")
    ws['I9'] = "ch no."
    ws['I9'].alignment = center_align

    # Date
    ws.merge_cells('G10:H10')
    ws['G10'] = "date:"
    ws['G10'].font = bold_font
    ws['G10'].alignment = center_align
    ws.merge_cells("I10:J10")
    ws["I10"] = "dd.mm.yy"
    ws["I10"].alignment = center_align

    ws.merge_cells('G11:J11')
    ws.merge_cells('A12:J12')
    
    # Item table
    headers = ["S. NO.", "ITEM NAME", "HSN", "PIECES", "AMOUNT"]
    col_positions = ["A", "B", "F", "H", "I"]
    ws.merge_cells('B13:E13')
    ws.merge_cells('F13:G13')
    ws.merge_cells('I13:J13')
    for col, header in zip(col_positions, headers):
        ws[f"{col}13"].value = header
        ws[f"{col}13"].font = bold_font
        ws[f"{col}13"].alignment = center_align
        ws[f"{col}13"].border = thin_border

    for i in range(14,21):
        ws.merge_cells(f'B{i}:E{i}')
        ws.merge_cells(f'F{i}:G{i}')
        ws.merge_cells(f'I{i}:J{i}')

    for i in range(21,27):
        ws.merge_cells(f'F{i}:G{i}')
        ws.merge_cells(f'I{i}:J{i}')

    # Footer
    ws["F21"].value = "TOTAL"
    ws["F22"].value = "DISCOUNT"
    ws["F23"].value = "GR"
    ws["F24"].value = "GST"
    ws["f25"].value = "NET AMOUNT"
    ws["f25"].font = Font(bold=True)
    ws["f21"].font = Font(bold=True)
    
    # OTHER Party goods
    ws.row_dimensions[28].height = 25
    ws['B28'] = 'other party goods'
    ws['B28'].font = Font(bold=True, size=16)

    # All cell capital
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):  # Check if the cell contains a string
                cell.value = cell.value.upper()

    #calling apply_borders function
    apply_borders(ws)

# Prompt user for input
def get_user_input(prompt, is_numeric=False):
    """Prompt user for input and validate if required."""
    while True:
        try:
            user_input = input(prompt)
            if is_numeric:
                return int(user_input)
            return user_input
        except ValueError:
            print("Invalid input. Please try again.")

# Function to generate transport challan and save it
def generate_challan(company_data, transport_data,contact_no, company_name, transport_name, items_data, discount, gst, date, challan_number,No_of_Other_Party_Goods,Amount_of_Other_Party_Goods):
    wb = Workbook()
    ws = wb.active
    format_ws(ws)

    company_details = company_data.get(company_name, {})
    transport_details = transport_data.get(transport_name, {})

    # Populate company and transport details
    ws["A2"].value = company_name
    ws["A3"].value = company_details.get("address1", "").upper()
    ws["A4"].value = company_details.get("address2", "").upper()
    ws["A5"].value = "GST:- " + company_details.get("gst", "")
    ws["H6"].value = contact_no

    ws["A9"].value = transport_name
    ws["A10"].value = transport_details.get("station", "")
    ws["A11"].value = "GST:- " + transport_details.get("gst", "")

    
    ws["D7"].value = transport_details.get("Way", "")
    ws["I10"].value = date
    ws["I9"].value = challan_number
    ws["I32"].value = company_name
    ws["I29"].font = Font(bold=True)

    row = 14
    total_amount = 0
    total_pieces = 0
    for item in items_data:
        item_name, hsn, pieces, amount = item
        total_amount += amount
        total_pieces += pieces

        ws[f"A{row}"].value = row - 13
        ws[f"B{row}"].value = item_name
        ws[f"F{row}"].value = hsn
        ws[f"H{row}"].value = pieces
        ws[f"I{row}"].value = amount

        row += 1

    # Footer totals
    ws["H21"].value = total_pieces
    ws["I21"].value = total_amount
    ws["H22"].value = discount 
    ws["H24"].value = gst
    ws["I25"].value = total_amount - discount + gst
    ws["H25"].font = Font(bold=True)
    ws["I25"].font = Font(bold=True)
    ws["I21"].font = Font(bold=True)

    ws["A28"].value = No_of_Other_Party_Goods
    ws['A28'].font = Font(bold=True, size=16)
    ws["F28"].value = Amount_of_Other_Party_Goods
    ws["F28"].font = Font(bold=True, size=16)

    company_name_underscore = company_name.replace(" ", "_")
    transport_name_underscore = transport_name.replace(" ", "_")
    date_underscore = date.replace(".", "_")

    initialize_counter()

    counter = get_next_counter()

    directory = "generated_challans"
    os.makedirs(directory, exist_ok=True)


    file_name = f"transport_challan_{company_name_underscore}_{transport_name_underscore}_{date_underscore}_{counter}.xlsx"
    filename=os.path.join(directory, file_name)
    wb.save(filename)
    os.startfile(filename)
    print(f"Challan saved as {filename}")

# Main function
def main():
    # Load company and transport data
    company_data = load_data_file(COMPANY_DATA_FILE)
    transport_data = load_data_file(TRANSPORT_DATA_FILE)

    # Get or add company details
    company_name = get_user_input("Enter company name: ").upper()
    if company_name not in company_data:
        print(f"Company {company_name} not found. Adding new company.")
        address1 = get_user_input("Enter company address(1): ")
        gst = get_user_input("Enter GSTIN: ")
        address2 = get_user_input("Enter contact address(2): ")
        company_data[company_name] = {"address1": address1, "address2": address2, "gst": gst}
        save_data_file(COMPANY_DATA_FILE, company_data)

    # Get or add transport details
    transport_name = get_user_input("Enter transport name: ").upper()
    if transport_name not in transport_data:
        print(f"Transport {transport_name} not found. Adding new transport.")
        station = get_user_input("Enter station: ").upper()
        gst = get_user_input("Enter transport gst: ")
        transport = get_user_input("Enter transport: ").upper()
        transport_data[transport_name] = {"station": station, "gst": gst, "way": transport}
        save_data_file(TRANSPORT_DATA_FILE, transport_data)

    # Collect item details from user
    items_data = []
    while True:
        item_name = get_user_input("Enter item name (or leave blank to finish): ")
        if not item_name:
            break
        hsn = get_user_input(f"Enter HSN for {item_name}: ")
        pieces = get_user_input(f"Enter number of pieces for {item_name}: ", is_numeric=True)
        amount = get_user_input(f"Enter amount for {item_name}: ", is_numeric=True)
        items_data.append((item_name, hsn, pieces, amount))

    # Collect financial details
    discount = get_user_input("Enter discount: ", is_numeric=True)
    gst = get_user_input("Enter GST: ", is_numeric=True)
    date = get_user_input("Enter date (DD.MM.YY): ")
    
    # Generate the challan
    generate_challan(company_data, transport_data,9375290850, company_name, transport_name, items_data, discount, gst, date, 123,34,13907)


if __name__ == "__main__":
    main()
